from unittest import TestCase

from typing import List
from pathlib import Path
import os

from openpyxl import load_workbook
import pandas as pd
import pantab

from config import PROJECT_BASE_DIR, DATA_BASE_DIR, logger

from helpers.pandas_helpers import dataframe_info
from helpers.file_helpers import (
    list_files,
    write_dataframes_excel,
    make_csv_from_parquet,
    make_csv_from_parquet_dir,
    make_hyper_from_parquet,
    make_hyper_from_parquet_dir,
)


class TestFileHelpers(TestCase):

    @classmethod
    def setUpClass(cls):
        # create a dataframe
        cls.df = pd.DataFrame({"a": [1, 2, 3], "b": [4, 5, 6]})
        fp = Path(DATA_BASE_DIR, "tests", "test_make_csv_from_parquet_dir.parquet")

        # make the parent directory if it doesn't exist
        fp.parent.mkdir(parents=True, exist_ok=True)
        cls.df.to_parquet(fp)

        # make 2 parquet files in directory
        cls.parquet_dir = Path(DATA_BASE_DIR, "tests", "helpers")
        cls.parquet_dir.mkdir(parents=True, exist_ok=True)
        cls.parquet_fp = Path(cls.parquet_dir, "test_df.parquet")
        cls.parquet_fp1 = Path(cls.parquet_dir, "test_df1.parquet")
        cls.df.to_parquet(cls.parquet_fp)
        cls.df.to_parquet(cls.parquet_fp1)

        # make an excel file in directory
        excel_fp = Path(cls.parquet_dir, "test_df.xlsx")
        cls.df.to_excel(excel_fp, index=False)
        super().setUpClass()
        logger.debug("TestFileHelpers.setUpClass")

    def setUp(self):
        pass

    def test_list_files(self):
        # creates 3 files in test dir and one directory
        sub_dir = Path(self.parquet_dir, "test_list_sub_dir")
        sub_dir.mkdir(parents=True, exist_ok=True)
        test_files = ["test1.csv", "test2.csv", "test3.csv"]
        for file in test_files:
            Path(sub_dir, file).touch()

        # check that there are 3 files in list
        files = list_files(sub_dir)
        self.assertEqual(len(files), 3, msg=f"files: {files}")
        for file in test_files:
            self.assertIn(file, files)

        # remove the files
        for file in test_files:
            os.remove(Path(sub_dir, file))

    def test_write_dataframes_excel(self):
        """test write_dataframes_excel"""
        # create a dataframe
        df = pd.DataFrame({"a": [1, 2, 3], "b": [4, 5, 6]})
        # create a test of dataframes with one non-dataframe
        dataframes = {
            "test": df,
            "test1": df,
            "test2": "test",
        }
        fp = Path(DATA_BASE_DIR, "tests", "test_write_dataframes_excel.xlsx")
        sheets = write_dataframes_excel(fp, dataframes)

        # check that the sheets are correct
        self.assertEqual(set(sheets), {"test", "test1"}, msg=f"sheets: {sheets}")
        # check that the file exists and has the correct sheets
        self.assertTrue(fp.exists())
        wb = load_workbook(filename=fp)
        wb_sheets = wb.sheetnames
        self.assertEqual(set(wb_sheets), set(sheets), msg=f"wb_sheets: {wb_sheets}")

        # remove the file
        os.remove(fp)

    def test_make_csv_from_parquet(self):
        """test make_csv_from_parquet"""
        # create a dataframe and save as parquet file
        df = pd.DataFrame({"a": [1, 2, 3], "b": [4, 5, 6]})
        fp = Path(DATA_BASE_DIR, "tests", "test_make_csv_from_parquet.parquet")
        df.to_parquet(fp)

        # make a csv from the parquet
        csv_fp = make_csv_from_parquet(fp)
        self.assertTrue(csv_fp.exists())

        # check that the csv has the same data as the parquet and ends in '.csv'
        csv_df = pd.read_csv(csv_fp)
        self.assertTrue(df.equals(csv_df))
        self.assertTrue(csv_fp.suffix == ".csv")

        # test with given csv file path
        fp_csv_test = Path(
            DATA_BASE_DIR, "tests", "test_make_csv_from_parquet_with_path.csv"
        )
        fp_test = make_csv_from_parquet(fp, fp_csv_test)
        self.assertTrue(fp_test.exists())

        # remove the file
        os.remove(fp_test)  #  with given csv file path
        os.remove(fp)  # original parquet file
        os.remove(csv_fp)  # without given csv file path

    def test_make_csv_from_parquet_dir(self):
        """test make_csv_from_parquet_dir"""

        # make a csv from the parquet
        csv_fps = make_csv_from_parquet_dir(self.parquet_dir)

        # should return 2 csv files
        self.assertEqual(len(csv_fps), 2)

        # check that the csv has the same data as the parquet and ends in '.csv'
        for csv_fp in csv_fps:
            csv_df = pd.read_csv(csv_fp)
            self.assertTrue(self.df.equals(csv_df))
            self.assertTrue(csv_fp.suffix == ".csv")

        # test with csv file path
        csv_dir = Path(DATA_BASE_DIR, "tests", "helpers", "test_sub_dir")
        csv_fps = make_csv_from_parquet_dir(self.parquet_dir, csv_dir)

        # should return 2 csv files
        self.assertEqual(len(csv_fps), 2)

        # check that the csv exists
        for csv_fp in csv_fps:
            self.assertTrue(csv_fp.exists())

    def test_make_hyper_from_parquet(self):
        """test make_hyper_from_parquet"""

        # make a hyper from the parquet
        hyper_fp, table_name = make_hyper_from_parquet(self.parquet_fp)
        # logger.debug(f"hyper_fp: {hyper_fp}, table_name: {table_name}")
        self.assertTrue(hyper_fp.exists())

        # check that the hyper has the same data as the parquet and ends in '.hyper'
        hyper_df = pantab.frame_from_hyper(hyper_fp, table=table_name)
        # convert all columns to int64 for comparison since hyper returns all columns as int64[pyarrow]
        hyper_df = hyper_df.astype("int64")
        # logger.debug(f"hyper_df: \n{dataframe_info(hyper_df)}")
        # logger.debug(f"Data read from Hyper file: \n{hyper_df}")
        # logger.debug(f"df.info: \n{dataframe_info(df)}")
        # logger.debug(f"df: \n{df}")
        self.assertTrue(self.df.equals(hyper_df))
        self.assertTrue(hyper_fp.suffix == ".hyper")

    def test_make_hyper_from_parquet_dir(self):
        """test make_hyper_from_parquet_dir"""

        hyper_fp0 = Path(self.parquet_dir, "test_df.hyper")

        # make a hyper from the parquet
        res = make_hyper_from_parquet_dir(self.parquet_dir, hyper_fp0)
        hyper_fp = res["hyper_file"]
        tables = res["tables"]

        # should return 2 hyper files
        self.assertEqual(len(tables), 2)

        # read in the hyper file and check it has 2 tables and each table has the same data as the parquet
        dfs = pantab.frames_from_hyper(hyper_fp)
        for table, df in dfs.items():
            # convert all columns to int64 for comparison since hyper returns all columns as int64[pyarrow]
            df = df.astype("int64")
            self.assertTrue(self.df.equals(df))

    @classmethod
    def tearDownClass(cls):

        # remove the files
        test_dir = Path(PROJECT_BASE_DIR, "datasets", "tests", "helpers")
        test_dir_files = list(test_dir.glob("*"))
        # filter for files only
        test_dir_files = [file for file in test_dir_files if file.is_file()]
        # logger.debug(f"test_dir_files: {test_dir_files}")
        # remove the files in the test_dir
        for file in test_dir_files:
            # logger.debug(f"file: {file}")
            os.remove(Path(test_dir, file))
        test_dir2 = Path(PROJECT_BASE_DIR, "datasets", "tests", "helpers", "sub_dir")
        test_dir_files2 = list(test_dir2.glob("*"))
        # filter for files only
        test_dir_files2 = [file for file in test_dir_files2 if file.is_file()]
        for file in test_dir_files2:
            os.remove(Path(test_dir2, file))
